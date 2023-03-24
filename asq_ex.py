import sys
import re
import fitz
import pandas as pd
from openpyxl import load_workbook

# Define the regular expression for matching the lines we're interested in
pattern = re.compile(r'(\d+)\.\s+(.+)\s+(Definitely disagree|Definitely agree|Slightly agree|Slightly disagree)')

# Check that two filenames were provided as command line arguments
if len(sys.argv) < 3:
    print("Usage: python extract_lines.py <pdf> <excel>")
    sys.exit(1)

# Open the PDF file and read its contents
pdf_filename = sys.argv[1]
with fitz.open(pdf_filename) as doc:
    page_count = doc.page_count
    lines = []
    for i in range(page_count):
        page = doc.load_page(i)
        phrases = ['Definitely agree', 'Definitely disagree', 'Slightly agree', 'Slightly disagree']
        text = page.get_text().replace('\n', ' ')  # Remove all newlines which break the regex
        for phrase in phrases:
            text = text.replace(phrase, phrase + '\n')
        for line in text.split('\n'):  # Split at every question number
            match = pattern.search(line)
            if match:
                lines.append(list(match.groups()))

# Create a pandas dataframe with the extracted data
df = pd.DataFrame(lines, columns=['Question Number', 'Question', 'Answer'])

# Write the dataframe to an Excel file
output_filename = 'output.xlsx'
df.to_excel(output_filename, index=False)

# Load the input Excel files
input_file_1 = output_filename
input_file_2 = sys.argv[2]
wb_1 = load_workbook(input_file_1)
wb_2 = load_workbook(input_file_2)

# Select the active worksheet in each workbook
ws_1 = wb_1.active
ws_2 = wb_2.active

# Write the answers to column D in both worksheets
for i, row in enumerate(ws_1.iter_rows(min_row=2, max_col=3), start=2):
    question_num = row[0].value
    answer = df.loc[df['Question Number'] == question_num, 'Answer'].iloc[0]
    ws_2.cell(row=i, column=4, value=answer)

# Save the updated workbooks
wb_2.save(input_file_2)
